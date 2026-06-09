from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from datetime import datetime, date
from typing import List, Optional
from pydantic import BaseModel
from uuid import UUID
import random, string

from app.core.database import get_db
from app.core.security import get_current_user, require_roles
from app.models.models import (
    ProcurementOrder, ProcurementItem, ProcurementStatus,
    Product, User, Notification, StockMovement, MovementType
)

router = APIRouter()


def gen_order_number():
    suffix = ''.join(random.choices(string.ascii_uppercase + string.digits, k=5))
    return f"ZK-{datetime.now().strftime('%y%m%d')}-{suffix}"


# ─── SCHEMAS ──────────────────────────────────────────
class ItemInput(BaseModel):
    product_id: UUID
    quantity: float
    notes: Optional[str] = None


class CreateOrder(BaseModel):
    items: List[ItemInput]
    notes: Optional[str] = None
    # Bozorchi to'g'ridan ochsa - narx ham bo'lishi mumkin
    direct_buy: bool = False
    prices: Optional[dict] = None   # {product_id: price}


class BuyerItemUpdate(BaseModel):
    item_id: Optional[UUID] = None   # mavjud item
    product_id: Optional[UUID] = None  # yangi qo'shilgan qator
    bought_qty: float
    price: Optional[float] = None


class BuyerConfirm(BaseModel):
    items: List[BuyerItemUpdate]
    notes: Optional[str] = None


class ReceiveItem(BaseModel):
    item_id: UUID
    received_qty: float
    damaged_qty: float = 0
    actual_price: Optional[float] = None
    checked: bool = False


class ReceiveInput(BaseModel):
    items: List[ReceiveItem]
    notes: Optional[str] = None


# ─── NOTIFICATION HELPER ──────────────────────────────
async def notify_role(db, role, title, message, order_id=None):
    users_r = await db.execute(select(User).where(User.role == role, User.is_active == True))
    for u in users_r.scalars().all():
        db.add(Notification(
            user_id=u.id, type="workflow", title=title, message=message,
            data={"order_id": order_id, "for_role": role} if order_id else {"for_role": role},
            is_read=False, sent_to_telegram=False,
        ))


def order_dict(o):
    return {
        "id": str(o.id),
        "order_number": o.order_number,
        "status": o.status.value if hasattr(o.status, 'value') else str(o.status),
        "total_estimated_cost": float(o.total_estimated_cost or 0),
        "total_actual_cost": float(o.total_actual_cost or 0),
        "buyer_notes": o.buyer_notes,
        "warehouse_notes": o.warehouse_notes,
        "created_at": o.created_at.isoformat() if o.created_at else None,
        "completed_at": o.completed_at.isoformat() if o.completed_at else None,
    }


async def get_product_stock(db, product_id):
    r = await db.execute(select(Product.current_stock).where(Product.id == product_id))
    v = r.scalar()
    return float(v) if v is not None else 0


# ─── 1. ZAKAZ YARATISH (Sotuvchi YOKI Bozorchi) ──────
@router.post("/")
async def create_order(
    data: CreateOrder,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles("seller", "buyer", "admin"))
):
    if not data.items:
        raise HTTPException(status_code=400, detail="Kamida bitta mahsulot kerak")

    role = current_user.role.value if hasattr(current_user.role, 'value') else current_user.role
    prices = data.prices or {}

    # Bozorchi to'g'ridan ochsa - darrov buyer_confirmed (omborchiga)
    is_buyer_direct = (role == "buyer") and data.direct_buy

    order = ProcurementOrder(
        order_number=gen_order_number(),
        status=ProcurementStatus.buyer_confirmed if is_buyer_direct else ProcurementStatus.ai_generated,
        buyer_notes=data.notes,
        created_at=datetime.utcnow(),
    )
    if is_buyer_direct:
        order.buyer_id = current_user.id
        order.buyer_confirmed_at = datetime.utcnow()
    db.add(order)
    await db.flush()

    total = 0.0
    for it in data.items:
        price = float(prices.get(str(it.product_id), 0)) if is_buyer_direct else None
        if is_buyer_direct and price:
            total += it.quantity * price
        db.add(ProcurementItem(
            order_id=order.id,
            product_id=it.product_id,
            ai_recommended_qty=it.quantity,
            buyer_ordered_qty=it.quantity if is_buyer_direct else None,
            estimated_price=price,
            notes=it.notes,
        ))

    if is_buyer_direct:
        order.total_estimated_cost = total
        await notify_role(db, "warehouse_manager", "📦 Tasdiqlash kerak", f"Bozorchidan to'g'ridan zakaz: {order.order_number}", str(order.id))
        msg = "Zakaz omborchiga yuborildi"
    else:
        await notify_role(db, "buyer", "🛒 Yangi zakaz", f"Sotuvchidan yangi zakaz: {order.order_number}", str(order.id))
        msg = "Zakaz bozorchiga yuborildi"

    await db.commit()
    return {"id": str(order.id), "order_number": order.order_number, "message": msg}


# ─── 2. BOZORCHI: TASDIQLASH (+yangi qator) ──────────
@router.post("/{order_id}/buyer-confirm")
async def buyer_confirm(
    order_id: UUID,
    data: BuyerConfirm,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles("buyer", "admin"))
):
    r = await db.execute(select(ProcurementOrder).where(ProcurementOrder.id == order_id))
    order = r.scalar_one_or_none()
    if not order:
        raise HTTPException(status_code=404, detail="Zakaz topilmadi")

    total = 0.0
    for upd in data.items:
        if upd.item_id:
            ir = await db.execute(select(ProcurementItem).where(ProcurementItem.id == upd.item_id))
            item = ir.scalar_one_or_none()
            if item:
                item.buyer_ordered_qty = upd.bought_qty
                item.estimated_price = upd.price
                if upd.price:
                    total += upd.bought_qty * upd.price
        elif upd.product_id:
            # Yangi qator qo'shilgan
            db.add(ProcurementItem(
                order_id=order.id,
                product_id=upd.product_id,
                ai_recommended_qty=0,
                buyer_ordered_qty=upd.bought_qty,
                estimated_price=upd.price,
            ))
            if upd.price:
                total += upd.bought_qty * upd.price

    order.status = ProcurementStatus.buyer_confirmed
    order.buyer_id = current_user.id
    order.buyer_notes = data.notes
    order.total_estimated_cost = total
    order.buyer_confirmed_at = datetime.utcnow()
    order.warehouse_notes = None  # eski rad sababini tozalaymiz

    await notify_role(db, "warehouse_manager", "📦 Tasdiqlash kerak", f"Bozorchidan zakaz: {order.order_number}", str(order.id))
    await db.commit()
    return {"message": "Zakaz omborchiga yuborildi", "status": order.status.value}


# ─── 3. OMBORCHI: TASDIQLASH / RAD ────────────────────
@router.patch("/{order_id}/approve")
async def approve_order(
    order_id: UUID,
    notes: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles("warehouse_manager", "admin"))
):
    r = await db.execute(select(ProcurementOrder).where(ProcurementOrder.id == order_id))
    order = r.scalar_one_or_none()
    if not order:
        raise HTTPException(status_code=404, detail="Zakaz topilmadi")
    order.status = ProcurementStatus.warehouse_approved
    order.warehouse_manager_id = current_user.id
    order.warehouse_approved_at = datetime.utcnow()
    await notify_role(db, "goods_receiver", "📥 Tovar kutilmoqda", f"Tasdiqlangan zakaz: {order.order_number}", str(order.id))
    await db.commit()
    return {"message": "Tovarovedga yuborildi", "status": order.status.value}


@router.patch("/{order_id}/reject")
async def reject_order(
    order_id: UUID,
    notes: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles("warehouse_manager", "admin"))
):
    """Omborchi rad etadi - bozorchiga qaytadi (sabab bilan)"""
    r = await db.execute(select(ProcurementOrder).where(ProcurementOrder.id == order_id))
    order = r.scalar_one_or_none()
    if not order:
        raise HTTPException(status_code=404, detail="Zakaz topilmadi")
    # Rad etilganda bozorchiga qaytadi (rejected emas, qayta tahrirlash uchun)
    order.status = ProcurementStatus.rejected
    order.warehouse_manager_id = current_user.id
    order.warehouse_notes = notes or "Sabab ko'rsatilmagan"
    await notify_role(db, "buyer", "❌ Zakaz qaytarildi", f"{order.order_number}: {notes or 'tuzatish kerak'}", str(order.id))
    await db.commit()
    return {"message": "Bozorchiga qaytarildi", "status": order.status.value}


# ─── 4. TOVAROVED: QABUL ──────────────────────────────
@router.post("/{order_id}/receive")
async def receive_goods(
    order_id: UUID,
    data: ReceiveInput,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles("goods_receiver", "admin"))
):
    r = await db.execute(select(ProcurementOrder).where(ProcurementOrder.id == order_id))
    order = r.scalar_one_or_none()
    if not order:
        raise HTTPException(status_code=404, detail="Zakaz topilmadi")

    total_actual = 0.0
    for ri in data.items:
        ir = await db.execute(select(ProcurementItem).where(ProcurementItem.id == ri.item_id))
        item = ir.scalar_one_or_none()
        if not item:
            continue
        item.received_qty = ri.received_qty
        item.damaged_qty = ri.damaged_qty
        item.actual_price = ri.actual_price
        if ri.actual_price:
            total_actual += ri.received_qty * ri.actual_price

        pr = await db.execute(select(Product).where(Product.id == item.product_id))
        product = pr.scalar_one_or_none()
        if product:
            net = ri.received_qty - ri.damaged_qty
            before = float(product.current_stock)
            product.current_stock = before + net
            # Oxirgi xarid narxini saqlaymiz
            if ri.actual_price:
                product.last_purchase_price = ri.actual_price
                product.purchase_price = ri.actual_price
            db.add(StockMovement(
                product_id=product.id, movement_type=MovementType.incoming,
                quantity=net, stock_before=before, stock_after=before + net,
                reference_id=order.id, reference_type="procurement_receive",
                user_id=current_user.id, notes=f"Zakaz {order.order_number}"
            ))

    order.status = ProcurementStatus.completed
    order.receiver_id = current_user.id
    order.total_actual_cost = total_actual
    order.completed_at = datetime.utcnow()
    await notify_role(db, "buyer", "✅ Tovar qabul qilindi", f"{order.order_number} yakunlandi", str(order.id))
    await db.commit()
    return {"message": "Tovar qabul qilindi, ombor yangilandi", "status": order.status.value}


# ─── RO'YXAT VA TAFSILOT ──────────────────────────────
@router.get("/")
async def list_orders(
    status: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = select(ProcurementOrder).order_by(ProcurementOrder.created_at.desc())
    if status:
        try:
            query = query.where(ProcurementOrder.status == ProcurementStatus(status))
        except ValueError:
            pass
    r = await db.execute(query)
    return [order_dict(o) for o in r.scalars().all()]


@router.get("/{order_id}")
async def get_order(
    order_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    r = await db.execute(select(ProcurementOrder).where(ProcurementOrder.id == order_id))
    order = r.scalar_one_or_none()
    if not order:
        raise HTTPException(status_code=404, detail="Zakaz topilmadi")

    items_r = await db.execute(
        select(ProcurementItem, Product.name, Product.unit, Product.current_stock,
               Product.last_purchase_price, Product.product_code)
        .join(Product, ProcurementItem.product_id == Product.id)
        .where(ProcurementItem.order_id == order_id)
    )
    d = order_dict(order)
    d["items"] = [
        {
            "id": str(it.id),
            "product_id": str(it.product_id),
            "product_name": name,
            "product_code": code or "",
            "unit": unit,
            "current_stock": float(stock or 0),           # oldingi/joriy ostatka
            "last_purchase_price": float(last_price or 0),  # oxirgi narx
            "requested_qty": float(it.ai_recommended_qty or 0),
            "bought_qty": float(it.buyer_ordered_qty or 0),
            "received_qty": float(it.received_qty or 0),
            "damaged_qty": float(it.damaged_qty or 0),
            "price": float(it.estimated_price or 0),
            "actual_price": float(it.actual_price or 0),
            "notes": it.notes,
        }
        for it, name, unit, stock, last_price, code in items_r.all()
    ]
    return d


# ─── EXCEL EXPORT (Nakladnoy) ─────────────────────────
@router.get("/{order_id}/excel")
async def export_excel(
    order_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    from fastapi.responses import StreamingResponse
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    r = await db.execute(select(ProcurementOrder).where(ProcurementOrder.id == order_id))
    order = r.scalar_one_or_none()
    if not order:
        raise HTTPException(status_code=404, detail="Zakaz topilmadi")

    items_r = await db.execute(
        select(ProcurementItem, Product.name, Product.unit, Product.product_code)
        .join(Product, ProcurementItem.product_id == Product.id)
        .where(ProcurementItem.order_id == order_id)
    )
    items = items_r.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Nakladnoy"
    bold = Font(bold=True)
    fill = PatternFill(start_color="3FB950", end_color="3FB950", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:H1")
    ws["A1"] = "BUYUK PREMIUM - NAKLADNOY"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = center
    ws["A2"] = f"Zakaz: {order.order_number}"
    ws["A2"].font = bold
    ws["A3"] = f"Sana: {order.created_at.strftime('%d.%m.%Y') if order.created_at else ''}"

    headers = ["№", "Kod", "Mahsulot", "Birlik", "Olingan", "Qabul", "Narx", "Summa"]
    row = 5
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill
        c.alignment = center
        c.border = border

    total = 0
    for i, (it, name, unit, code) in enumerate(items, 1):
        row += 1
        received = float(it.received_qty or 0)
        price = float(it.actual_price or it.estimated_price or 0)
        summa = received * price
        total += summa
        vals = [i, code or "", name, unit, float(it.buyer_ordered_qty or 0), received, price, summa]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.border = border
            if col >= 5: c.alignment = center

    row += 1
    ws.cell(row=row, column=7, value="JAMI:").font = bold
    ws.cell(row=row, column=8, value=total).font = bold

    for col, w in enumerate([5, 10, 25, 8, 10, 10, 12, 15], 1):
        ws.column_dimensions[chr(64 + col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=nakladnoy_{order.order_number}.xlsx"}
    )
